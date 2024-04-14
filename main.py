import sys
import datetime
import wikipedia

import webbrowser
import os
import smtplib
import openpyxl
import subprocess
import pywhatkit
import PyQt5
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QTextEdit, QLabel
import speech_recognition as sr
import pyttsx3
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pyjokes



engine = pyttsx3.init('sapi5')


def speak(audio):
    engine.say(audio)
    engine.runAndWait()


def attendancetracker():
    # Load the excel file containing attendance data
    wb = openpyxl.load_workbook(r'C:\Users\harsh\Downloads\Attendance_tracker24 (3).xlsx')
    sheet = wb['Sheet1']

    # Define a list to store student email addresses
    emails = []

    # Define the sender email address and password
    sender_email = "Email"
    password = "Password"

    # Connect to the Gmail SMTP server
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.ehlo()
    server.starttls()
    server.ehlo()
    server.login(sender_email, password)

    # Loop through the list of student email addresses
    # for email in emails:
    # Create a new MIME message
    for row in range(2, sheet.max_row + 1):
        email = sheet.cell(row=row, column=2).value

        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = email
        msg['Subject'] = "Attendance Percentage"

        # Get the attendance percentage for the current student

        attendance_column = sheet.cell(row=row, column=3)
        attendance_percentage = sheet.cell(row=row, column=3).value
        row1 = (attendance_percentage / 44) * 100

        # Create the email body
        body = "Your attendance percentage in OS is: " + str(row1) + "%\n"
        msg.attach(MIMEText(body, 'plain'))

        # row = sheet.max_row
        attendance_column = sheet.cell(row=row, column=4)
        attendance_percentage = sheet.cell(row=row, column=4).value
        row2 = (attendance_percentage / 44) * 100

        # Create the email body
        body = "Your attendance percentage in Java is: " + str(row2) + "%\n"
        msg.attach(MIMEText(body, 'plain'))

        # row = sheet.max_row
        attendance_column = sheet.cell(row=row, column=5)
        attendance_percentage = sheet.cell(row=row, column=5).value
        row3 = (attendance_percentage / 44) * 100

        # Create the email body
        body = "Your attendance percentage in DS is: " + str(row3) + "%\n"
        msg.attach(MIMEText(body, 'plain'))

        # row = sheet.max_row
        attendance_column = sheet.cell(row=row, column=6)
        attendance_percentage = sheet.cell(row=row, column=6).value
        row4 = (attendance_percentage / 44) * 100

        # Create the email body
        body = "Your attendance percentage in DBMS is: " + str(row4) + "%\n"
        msg.attach(MIMEText(body, 'plain'))

        # row = sheet.max_row
        attendance_column = sheet.cell(row=row, column=7)
        attendance_percentage = sheet.cell(row=row, column=7).value
        row5 = (attendance_percentage / 44) * 100

        # Create the email body
        body = "Your attendance percentage in CN is: " + str(row5) + "%\n"
        msg.attach(MIMEText(body, 'plain'))

        # Send the email
        server.sendmail(sender_email, email, msg.as_string())

    # Close the SMTP connection
    server.quit()

def note(text):
    date = datetime.datetime.now()
    file_name = str(date).replace(":", "-") + "-note.txt"

    with open(file_name, "w") as f:
        f.write(text)

    subprocess.Popen(["notepad.exe", file_name])

def wishMe():
    hour = datetime.datetime.now().hour
    if hour >= 0 and hour < 12:
        speak("Hello,Good Morning")
    elif hour >= 12 and hour < 18:
        speak("Hello,Good Afternoon")
    else:
        speak("Hello,Good Evening")

class VoiceAssistantWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Initialize the speech recognition and text-to-speech engines
        self.speech_recognizer = sr.Recognizer()
        self.speech_engine = pyttsx3.init()

        # Set up the GUI
        self.setWindowTitle('Desktop Assistant')
        self.setWindowIcon(PyQt5.QtGui.QIcon(r'C:\Users\harsh\Downloads\mic.png'))
        self.setGeometry(100, 100, 400, 400)

        # Add a button to start the voice recognition
        self.start_button = QPushButton('Start', self)
        self.start_button.setGeometry(10, 10, 50, 30)
        self.start_button.clicked.connect(self.start_voice_recognition)

        # Add a text box to display the recognized speech
        self.speech_textbox = QTextEdit(self)
        self.speech_textbox.setGeometry(10, 50, 380, 340)

        # Add a label to display the status of the voice recognition
        self.status_label = QLabel('Press the button to start voice recognition', self)
        self.status_label.setGeometry(70, 10, 300, 30)
        wishMe()

    def start_voice_recognition(self):
        # Change the status label
        self.status_label.setText('Listening...')

        # Start listening to the microphone
        with sr.Microphone() as source:
            audio = self.speech_recognizer.listen(source)

        # Try to recognize the speech
        try:
            recognized_text = self.speech_recognizer.recognize_google(audio)
            self.speech_textbox.append(recognized_text)
            self.status_label.setText('Recognized speech: ' + recognized_text)
        except sr.UnknownValueError:
            self.status_label.setText('Unable to recognize speech')
        except sr.RequestError as e:
            self.status_label.setText('Error: ' + str(e))

        query = recognized_text.lower()

        if 'wikipedia' in query:
            speak('Searching Wikipedia...')
            query = query.replace("wikipedia", "")
            results = wikipedia.summary(query, sentences=2)
            speak("According to Wikipedia")
            print(results)
            speak(results)

        elif 'open youtube' in query:
            webbrowser.open("youtube.com")


        elif 'open google' in query:
            webbrowser.open("google.com")

        elif 'open stackoverflow' in query:
            webbrowser.open("stackoverflow.com")


        elif 'the time' in query:
            strTime = datetime.datetime.now().strftime("%H:%M:%S")
            speak(f"Sir, the time is {strTime}")

        elif 'open vs code' in query:
            Path = r"C:\Users\harsh\AppData\Local\Programs\Microsoft VS Code\Code.exe"
            os.startfile(Path)

        elif 'open pycharm' in query:
            Path = r"C:\Program Files\JetBrains\PyCharm Community Edition 2021.3.3\bin\pycharm64.exe"
            os.startfile(Path)

        elif 'make a note' in query:
            statement = query.replace("make a note", "")
            note(statement)

        elif 'play' in query:
            song = query.replace('play', '')
            speak('playing' + song)
            pywhatkit.playonyt(song)

        elif 'send attendance report' in query:
            speak('please wait! sending mails to the students')
            attendancetracker()
            speak('Mails sent to the students')

        elif 'joke' in query:
            speak(pyjokes.get_joke())

        elif 'news' in query:
            webbrowser.open("https://timesofindia.indiatimes.com/")
            speak('Here are some headlines from the Times of India')

        elif 'open netflix' in query:
            webbrowser.open_new_tab("netflix.com/browse")
            speak("Netflix open now")

        elif 'open prime video' in query:
            webbrowser.open_new_tab("primevideo.com")
            speak("Amazon Prime Video open now")


        elif 'send email' in query:
            # Email account information
            email_address = "your id"
            email_password = "password"

            # Email recipient and message
            recipient = "recipient id"
            subject = "Voice Assistant Email"
            body = "This is a message sent from a voice assistant."

            # Create the message
            message = f"Subject: {subject}\n\n{body}"

            # Send the email
            with smtplib.SMTP("smtp.gmail.com", 587) as server:
                server.starttls()
                server.login(email_address, email_password)
                server.sendmail(email_address, recipient, message)

        # Speak the recognized text
        self.speech_engine.say(recognized_text)
        self.speech_engine.runAndWait()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = VoiceAssistantWindow()
    window.show()
    sys.exit(app.exec_())